# F16 Stage 5c：Output 卡產的東西 == Export 精靈產的東西（逐位元組）。
"""**這一份是刪掉 Export 精靈的許可證。**

使用者定調「Output 為真相，現有的 export 精靈請拿掉」。可是「換一條路」跟
「換一個結果」在畫面上長得一模一樣 —— 一份少了兩欄的 CSV、一份 topn 少寫了
幾顆的 KLARF、一疊沒有分數的 PNG，看起來都像是成功了。所以刪之前要先量：
**同一批結果，兩條路各產一次，逐位元組比對。**

比得起來的原因是 Output 卡**薄薄一層包在 `core/export/` 上** —— 演算法一行都
沒有重寫。所以這一份其實在證兩件事：

1. 那一層真的只是轉呼叫（沒有偷偷改預設值、漏送選項）；
2. `core/export/` 是**共用**的引擎，精靈走了它一行都不會少。

⚠ **精靈刪掉之後這一份不會跟著消失。** `_wizard_*` 幾支會改成直接叫
`core/export/`（那正是精靈今天做的事），所以它繼續守著同一件事：Output 卡
不准跟引擎的預設值漂開。

Excel 只比**格子**不比位元組：`.xlsx` 是一個 zip，而 zip 的檔頭裡有時間戳
（同一份內容連寫兩次的位元組本來就不同）—— 比位元組會變成一條每次都紅的測試。
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
sys.path.insert(0, str(REPO / "tools"))

import d4t.core.steps  # noqa: F401,E402 — 觸發卡片註冊
from d4t.core.export import (  # noqa: E402
    apply_writeback, overlay_filename, overlay_label, pick_overlay_results,
    render_overlay, write_csv, write_excel, write_png,
)
from d4t.core.ingest.dataset import load_dataset  # noqa: E402
from d4t.core.pipeline import run_batch, run_batch_steps, run_defect  # noqa: E402
from d4t.core.pipeline.recipe import Recipe, RecipeNode, ScoreSpec  # noqa: E402

KIND = "ebi_patch"
N = 6


@pytest.fixture(scope="module")
def lot(tmp_path_factory):
    from make_sample import generate
    return generate(str(tmp_path_factory.mktemp("parity")), n=N, seed=7)


def _recipe(step: str, params: dict) -> Recipe:
    """Load → Gray level → 一張 Output 卡。"""
    return Recipe(
        recipe_id="parity", routes={KIND: ["load", "glv", "out"]},
        nodes={
            "load": RecipeNode("load", "load_patch", {}),
            "glv": RecipeNode("glv", "glv_stats",
                              {"source": "test", "metrics": "glv_max,glv_mean"}),
            "out": RecipeNode("out", step, params),
        },
        score=ScoreSpec(expr="glv_max", threshold=1.0,
                        bins={"below": 0, "above": 1}))


def _both(lot, tmp_path, step, name, **params):
    """跑一次批次，回傳 ``(recipe, rows, dataset, 精靈的路徑, 卡片的路徑)``。

    兩邊各自 `load_dataset` 一次：`apply_writeback` 會在 doc 上寫格子，
    共用一份的話第二條路吃到的是第一條路改過的東西 —— 那樣比出來的「一樣」
    不算數。
    """
    wiz_dir, card_dir = tmp_path / "wizard", tmp_path / "card"
    wiz_dir.mkdir(exist_ok=True)
    card_dir.mkdir(exist_ok=True)
    key = "folder" if step == "output_image" else "path"
    recipe = _recipe(step, dict(params, **{key: str(card_dir / name)}))
    dataset = load_dataset(lot["klarf"])
    rows = run_batch(recipe, dataset, workers=1)
    assert len(rows) == N and all(r["ok"] for r in rows)
    return recipe, rows, dataset, wiz_dir / name, card_dir / name


def _run_card(recipe, dataset, rows):
    bctx = run_batch_steps(recipe, dataset, rows)
    assert not bctx.errors, bctx.errors
    return bctx


# --------------------------------------------------------------------------- #
# 1. CSV
# --------------------------------------------------------------------------- #
def test_the_csv_card_writes_the_same_bytes_as_the_wizard(lot, tmp_path):
    recipe, rows, dataset, wiz, card = _both(lot, tmp_path, "output_csv",
                                             "defects.csv")
    write_csv(rows, str(wiz))                       # 精靈那一行
    _run_card(recipe, dataset, rows)
    assert card.read_bytes() == wiz.read_bytes()
    assert len(card.read_text(encoding="utf-8-sig").splitlines()) == N + 1


# --------------------------------------------------------------------------- #
# 2. Excel（比格子不比位元組 —— 見模組 docstring）
# --------------------------------------------------------------------------- #
def _sheets(path):
    from openpyxl import load_workbook
    wb = load_workbook(str(path))
    return {ws.title: [list(r) for r in ws.iter_rows(values_only=True)]
            for ws in wb.worksheets}


def test_the_report_card_writes_the_same_workbook_as_the_wizard(lot, tmp_path):
    recipe, rows, dataset, wiz, card = _both(lot, tmp_path, "output_report",
                                             "report.xlsx")
    pytest.importorskip("openpyxl")
    write_excel(rows, str(wiz), recipe=recipe, ground_truth=None)
    _run_card(recipe, dataset, rows)
    a, b = _sheets(wiz), _sheets(card)
    assert list(a) == list(b), "工作表不一樣"
    for name in a:
        assert a[name] == b[name], "工作表 %s 的內容不一樣" % name


# --------------------------------------------------------------------------- #
# 3. KLARF —— 三種模式各一條
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("mode,params,opts", [
    ("annotate", {}, {}),
    ("topn", {"top_n": 3, "min_score": 0.0, "renumber": True,
              "include_annotations": True},
     {"n": 3, "min_score": 0.0, "renumber": True,
      "include_annotations": True}),
    ("topn", {"top_n": 0, "min_score": 1.0, "renumber": False,
              "include_annotations": False},
     {"n": 0, "min_score": 1.0, "renumber": False,
      "include_annotations": False}),
    ("inplace", {"class_col": "CLASSNUMBER"},
     {"class_col": "CLASSNUMBER", "size_scale": 1.0,
      "size_feature": "cd_x_px"}),
])
def test_the_klarf_card_writes_the_same_bytes_as_the_wizard(
        lot, tmp_path, mode, params, opts):
    """**三種模式各一條**，因為這個 bug 的形狀就是「只測了一種」：
    上一輪 topn 送錯了關鍵字，而只覆蓋 annotate 的測試整組是綠的。
    """
    sub = tmp_path / ("m_%s_%s" % (mode, params.get("top_n", "x")))
    sub.mkdir()
    recipe, rows, dataset, wiz, card = _both(lot, sub, "output_klarf",
                                             "out.001", mode=mode, **params)
    if mode == "inplace":
        # inplace 改的是**指到的那個檔** —— 兩邊各給一份原檔的副本。
        shutil.copyfile(lot["klarf"], wiz)
        shutil.copyfile(lot["klarf"], card)
    apply_writeback(load_dataset(lot["klarf"]).klarf, rows, mode, str(wiz),
                    **opts)
    _run_card(recipe, dataset, rows)
    assert card.exists(), "卡片什麼都沒寫出來（mode=%s）" % mode
    assert card.read_bytes() == wiz.read_bytes()


def test_inplace_with_nothing_named_leaves_the_file_alone(lot, tmp_path):
    """鐵則 6 的那一句在**卡片這條路上**也要成立：一格都不填 = 逐位元組相同。"""
    recipe, rows, dataset, _, card = _both(lot, tmp_path, "output_klarf",
                                           "same.001", mode="inplace")
    shutil.copyfile(lot["klarf"], card)
    _run_card(recipe, dataset, rows)
    assert card.read_bytes() == Path(lot["klarf"]).read_bytes()


# --------------------------------------------------------------------------- #
# 4. 疊圖 PNG
# --------------------------------------------------------------------------- #
def test_the_image_card_writes_the_same_pngs_as_the_wizard(lot, tmp_path):
    """精靈的疊圖有**兩件卡片以前沒有**的事，而兩件都看不出來：

    * 依分數由高到低取前 N（照 `rows` 的順序取的話拿到的是**檔案順序**上的
      前 N 顆）；
    * 左上角那一行有 `score=` 與 `bin=`（少了的話那疊 PNG 跟完整的長得一樣，
      而使用者正是拿它們來挑門檻的）。

    兩件都搬進了 `core/export/overlay.py`，所以這一條比得起來。
    """
    recipe, rows, dataset, wiz, card = _both(lot, tmp_path, "output_image",
                                             "png", limit=3)
    wiz.mkdir(parents=True, exist_ok=True)
    by_id = {str(it.defect_id): it for it in dataset.items}
    for row in pick_overlay_results(rows, 3):       # 精靈那幾行
        r = run_defect(recipe, by_id[str(row["defect_id"])], KIND,
                       keep_context=True)
        write_png(render_overlay(dict(r.context.images), dict(r.features),
                                 label=overlay_label(row)),
                  str(wiz / overlay_filename(row["defect_id"])))
    _run_card(recipe, dataset, rows)

    got = sorted(p.name for p in card.glob("*.png"))
    assert len(got) == 3
    assert got == sorted(p.name for p in wiz.glob("*.png"))
    for name in got:
        assert (card / name).read_bytes() == (wiz / name).read_bytes(), name
