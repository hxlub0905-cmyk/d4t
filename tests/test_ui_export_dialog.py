# ADEPT Studio 輸出精靈測試 — authored 2026-07-28 (M5-3).
"""``adept/ui/export_dialog.py``（KLARF 寫回 / 報表 / 疊圖）的離屏測試。

執行：``QT_QPA_PLATFORM=offscreen python3 -m pytest tests/test_ui_export_dialog.py -q``

**為什麼所有 Qt import 都是 lazy 的（別改回去）**

``tests/test_no_qt.py::test_no_qt_after_import`` 會檢查 ``sys.modules`` 裡沒有任何
PySide6 模組。pytest 先蒐集全部測試檔、再開始跑，所以只要這個檔案在**模組層**
``import PySide6``（或 import ``adept.ui.export_dialog``），蒐集階段就會把 Qt 塞進
``sys.modules``，那個守門測試就會紅 —— 即使它先跑。

因此：所有 Qt / ``adept.ui`` 的 import 都關在 :func:`_load_qt` 裡，由 module-scope
的 ``qapp`` fixture 呼叫，再用 ``globals().update(...)`` 注入本模組命名空間。
每個測試都必須（直接或間接）要求 ``qapp`` fixture，否則那些名字不存在。

這個檔案守的四條底線：

1. **控制項要反映真實的 KLARF**：有的欄位才能選，沒有的欄位變灰**並寫明原因**。
2. **「預覽變更」是硬性關卡**：沒預覽 → 「寫出」是灰的；設定一改 → 預覽作廢。
3. **無損**：inplace 什麼都不改 → 輸出檔與原檔**逐位元組相同**（鐵則 6）。
4. **錯誤是人話**：:class:`ExportError` 變成一句看得懂的訊息，不是 traceback。
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parent.parent
EXAMPLE_RECIPE = REPO / "examples" / "recipes" / "die_to_die_basic.json"

sys.path.insert(0, str(REPO / "tools"))
from make_sample import generate  # noqa: E402

from adept.core.export import feature_keys  # noqa: E402
from adept.core.ingest import klarf_core  # noqa: E402
from adept.core.ingest.dataset import load_dataset  # noqa: E402
from adept.core.pipeline import Recipe, run_batch  # noqa: E402

N = 8


def _load_qt() -> None:
    """把 Qt 與待測模組 import 進來，注入本模組的 globals（只在 fixture 裡呼叫）。"""
    from PySide6.QtCore import Qt  # noqa: F401
    from PySide6.QtWidgets import QApplication  # noqa: F401

    from adept.ui import export_dialog as ex_mod  # noqa: F401
    from adept.ui import theme as theme_mod  # noqa: F401

    globals().update(locals())


@pytest.fixture(scope="module")
def qapp():
    """離屏 QApplication（整個模組共用一個）+ 套用主題。"""
    _load_qt()
    app = QApplication.instance() or QApplication([sys.argv[0] if sys.argv else "t"])
    theme_mod.apply_theme(app)
    yield app
    app.processEvents()


@pytest.fixture(scope="module")
def synlot(tmp_path_factory):
    """合成 lot（8 顆 EBI patch defect）—— 與其他測試同一支產生器。"""
    return generate(str(tmp_path_factory.mktemp("export_lot")), n=N, seed=7)


@pytest.fixture(scope="module")
def ran(synlot):
    """真的跑一次 die-to-die 範例流程（不是手捏的假結果）。"""
    ds = load_dataset(synlot["klarf"])
    recipe = Recipe.load(str(EXAMPLE_RECIPE))
    results = run_batch(recipe, ds, workers=1)
    assert len(results) == N
    assert any(r["ok"] for r in results)
    return {"dataset": ds, "recipe": recipe, "results": results,
            "doc": ds.klarf, "klarf": synlot["klarf"]}


@pytest.fixture
def dlg(qapp, ran, tmp_path):
    """一個全新的對話框（每個測試各自一份，設定不互相污染）。"""
    d = ex_mod.ExportDialog(ran["results"], doc=ran["doc"],
                            dataset=ran["dataset"], recipe=ran["recipe"],
                            default_dir=str(tmp_path))
    yield d
    d.close()


# --------------------------------------------------------------------------- #
# 1. 三個模式的控制項都由「真實的欄位清單」長出來
# --------------------------------------------------------------------------- #
def test_controls_populate_from_real_klarf_columns(dlg, ran):
    assert dlg.mode() == "inplace"
    assert dlg.klarf_columns() == list(ran["doc"].defect_columns)

    # 這份合成 KLARF 只有 CLASSNUMBER，另外三欄都不存在
    assert "CLASSNUMBER" in dlg.klarf_columns()
    assert dlg.column_enabled("CLASSNUMBER") is True
    for name in ("ROUGHBINNUMBER", "FINEBINNUMBER", "DSIZE"):
        assert name not in dlg.klarf_columns()
        assert dlg.column_enabled(name) is False, name
        hint = dlg.column_hint(name)
        assert name in hint and "has no" in hint, hint       # 講清楚為什麼不能選
        assert dlg.column_control(name).isChecked() is False

    # annotate：特徵多選清單 = 這批結果真的量到的特徵
    assert dlg.feature_names() == feature_keys(ran["results"])
    assert "snr_max" in dlg.feature_names()
    assert dlg.selected_features() == []

    # topn：取幾顆 / 分數門檻 / 重新編號
    assert dlg.set_mode("topn") is True
    assert dlg.spin_topn.value() >= 1
    assert dlg.chk_renumber.isChecked() is True
    assert dlg.set_mode("nope") is False


def test_mode_pages_swap(dlg):
    dlg.show()
    for mode, page in (("inplace", dlg.page_inplace),
                       ("annotate", dlg.page_annotate),
                       ("topn", dlg.page_topn)):
        dlg.set_mode(mode)
        assert dlg.mode() == mode
        assert page.isVisible() is True
        others = [p for p in (dlg.page_inplace, dlg.page_annotate, dlg.page_topn)
                  if p is not page]
        assert all(p.isVisible() is False for p in others)


# --------------------------------------------------------------------------- #
# 2. 「預覽變更」是硬性關卡
# --------------------------------------------------------------------------- #
def test_preview_fills_plan_and_enables_write(dlg, tmp_path):
    assert dlg.btn_write.isEnabled() is False, "沒預覽就不准寫"
    assert dlg.plan_text() == ""

    dlg.set_output_path("klarf", str(tmp_path / "out.001"))
    dlg.chk_class_col.setChecked(True)
    plan = dlg.preview_plan()

    assert plan is not None
    assert plan.mode == "inplace"
    assert plan.columns_touched == ["CLASSNUMBER"]
    assert plan.n_rows_out == N
    assert plan.n_rows_changed > 0

    text = dlg.plan_text()
    assert text == dlg.plan_view.toPlainText()
    for piece in ("Mode: ", "Rows changed: %d defects" % plan.n_rows_changed,
                  "Existing columns touched: CLASSNUMBER", "Columns added: (none)",
                  "The output file will have %d defect rows" % plan.n_rows_out,
                  "Output health check:"):
        assert piece in text, piece
    assert "✓" in text                                  # 這份檔案健檢是乾淨的
    assert dlg.btn_write.isEnabled() is True

    # 任何設定被動過 → 計畫作廢、按鈕鎖回去（使用者看到的一定是這一版的後果）
    dlg.chk_class_col.setChecked(False)
    assert dlg.plan() is None
    assert dlg.plan_text() == ""
    assert dlg.btn_write.isEnabled() is False
    assert "press “Preview changes” again" in dlg.plan_view.toPlainText()


def test_write_without_preview_is_refused(dlg, tmp_path):
    dlg.set_output_path("klarf", str(tmp_path / "never.001"))
    assert dlg.run_export(sync=True) is None
    assert "Preview changes" in dlg.error_text()
    assert not (tmp_path / "never.001").exists()


def test_plan_with_lint_errors_surfaces_them(qapp, ran, tmp_path):
    """健檢有 error 的檔案 → 計畫書裡要看得到 ✗ 那幾行。"""
    broken_path = tmp_path / "broken.001"
    text = Path(ran["klarf"]).read_text(encoding="utf-8")
    broken_path.write_text(text.replace("EndOfFile;", ""), encoding="utf-8")
    doc = klarf_core.load(str(broken_path))

    d = ex_mod.ExportDialog(ran["results"], doc=doc, default_dir=str(tmp_path))
    try:
        d.set_output_path("klarf", str(tmp_path / "out_broken.001"))
        plan = d.preview_plan()
        assert plan is not None
        levels = [i.level for i in plan.issues]
        assert "error" in levels, levels
        text_out = d.plan_text()
        assert "✗" in text_out
        assert "EndOfFile" in text_out                   # 說出是哪個問題
    finally:
        d.close()


# --------------------------------------------------------------------------- #
# 3. 無損：inplace 什麼都不改 → 逐位元組相同
# --------------------------------------------------------------------------- #
def test_inplace_noop_write_is_byte_identical(dlg, ran, tmp_path):
    out = tmp_path / "noop.001"
    dlg.set_output_path("klarf", str(out))
    assert dlg.column_control("CLASSNUMBER").isChecked() is False

    plan = dlg.preview_plan()
    assert plan is not None and plan.n_rows_changed == 0
    assert "byte-for-byte identical" in dlg.plan_text()

    summary = dlg.run_export(sync=True)
    assert dlg.error_text() == ""
    assert summary is not None
    assert str(out) in summary["outputs"]

    assert out.read_bytes() == Path(ran["klarf"]).read_bytes()


def test_inplace_writes_classnumber(dlg, ran, tmp_path):
    out = tmp_path / "cls.001"
    dlg.set_output_path("klarf", str(out))
    dlg.chk_class_col.setChecked(True)
    assert dlg.preview_plan() is not None
    assert dlg.run_export(sync=True) is not None

    doc = klarf_core.load(str(out))
    j = doc.col_index("CLASSNUMBER")
    di = doc.col_index("DEFECTID")
    got = {row[di]: row[j] for row in doc.defects}
    for r in ran["results"]:
        if r.get("ok") and r.get("bin") is not None:
            assert got[str(r["defect_id"])] == str(int(r["bin"]))


def test_annotate_adds_score_class_and_chosen_features(dlg, tmp_path):
    out = tmp_path / "annot.001"
    dlg.set_mode("annotate")
    dlg.set_output_path("klarf", str(out))
    dlg.set_selected_features(["snr_max"])
    assert dlg.selected_features() == ["snr_max"]

    plan = dlg.preview_plan()
    assert plan is not None
    assert plan.columns_added == ["ADCSCORE", "ADCCLASS", "SNR_MAX"]
    assert "Columns added: ADCSCORE, ADCCLASS, SNR_MAX" in dlg.plan_text()

    assert dlg.run_export(sync=True) is not None
    doc = klarf_core.load(str(out))
    for col in ("ADCSCORE", "ADCCLASS", "SNR_MAX"):
        assert doc.col_index(col) >= 0, col
    # 影像參照不能壞：新欄位插在影像欄之前，IMAGELIST 仍在列尾
    assert doc.defect_image_map()["mode"] is not None
    assert len(doc.defects) == N


def test_topn_keeps_only_the_best(dlg, tmp_path):
    out = tmp_path / "top3.001"
    dlg.set_mode("topn")
    dlg.set_output_path("klarf", str(out))
    dlg.spin_topn.setValue(3)

    plan = dlg.preview_plan()
    assert plan is not None
    assert plan.n_rows_out == 3
    assert "The output file will have 3 defect rows" in dlg.plan_text()

    assert dlg.run_export(sync=True) is not None
    doc = klarf_core.load(str(out))
    assert len(doc.defects) == 3
    di = doc.col_index("DEFECTID")
    assert [row[di] for row in doc.defects] == ["1", "2", "3"]   # 重新編號


# --------------------------------------------------------------------------- #
# 4. 錯誤是人話
# --------------------------------------------------------------------------- #
def test_export_error_is_readable_not_traceback(qapp, ran, tmp_path):
    """對一份已經有 ADCSCORE 的 KLARF 再 annotate → 白話錯誤，不是 traceback。"""
    first = tmp_path / "once.001"
    d1 = ex_mod.ExportDialog(ran["results"], doc=ran["doc"], default_dir=str(tmp_path))
    try:
        d1.set_mode("annotate")
        d1.set_output_path("klarf", str(first))
        assert d1.preview_plan() is not None
        assert d1.run_export(sync=True) is not None
    finally:
        d1.close()

    d2 = ex_mod.ExportDialog(ran["results"], doc=klarf_core.load(str(first)),
                             default_dir=str(tmp_path))
    try:
        d2.set_mode("annotate")
        d2.set_output_path("klarf", str(tmp_path / "twice.001"))
        assert d2.preview_plan() is None
        assert d2.plan() is None
        assert d2.btn_write.isEnabled() is False
        msg = d2.error_text()
        assert "ADCSCORE" in msg and "Traceback" not in msg
        assert "✗" in d2.plan_view.toPlainText()
        assert not (tmp_path / "twice.001").exists()
    finally:
        d2.close()


# --------------------------------------------------------------------------- #
# 5. 報表與疊圖
# --------------------------------------------------------------------------- #
def test_csv_and_excel_checkboxes_produce_files(dlg, ran, tmp_path):
    csv_path = tmp_path / "features.csv"
    xlsx_path = tmp_path / "report.xlsx"
    dlg.chk_klarf.setChecked(False)          # 只出報表：不必先預覽
    dlg.chk_csv.setChecked(True)
    dlg.chk_excel.setChecked(True)
    dlg.set_output_path("csv", str(csv_path))
    dlg.set_output_path("excel", str(xlsx_path))
    assert dlg.btn_write.isEnabled() is True

    summary = dlg.run_export(sync=True)
    assert dlg.error_text() == ""
    assert summary is not None
    assert str(csv_path) in summary["outputs"]
    assert str(xlsx_path) in summary["outputs"]

    head = csv_path.read_text(encoding="utf-8-sig").splitlines()
    assert head[0].startswith("defect_id,ok,error,score,bin")
    assert len(head) == N + 1

    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path))
    assert wb.sheetnames == ["Summary", "Details", "Feature stats"]

    assert "CSV" in dlg.summary_text() or "csv" in dlg.summary_text().lower()


def test_overlay_writes_pngs_with_limit(dlg, tmp_path):
    out_dir = tmp_path / "ov"
    dlg.chk_klarf.setChecked(False)
    dlg.chk_overlay.setChecked(True)
    dlg.spin_overlay_limit.setValue(2)
    dlg.set_output_path("overlay", str(out_dir))

    summary = dlg.run_export(sync=True)
    assert dlg.error_text() == ""
    assert summary is not None and summary["n_overlays"] == 2

    pngs = sorted(p for p in os.listdir(str(out_dir)) if p.endswith(".png"))
    assert len(pngs) == 2
    assert all(p.startswith(ex_mod.OVERLAY_PREFIX) for p in pngs)
    assert (out_dir / pngs[0]).stat().st_size > 0


def test_nothing_selected_is_refused(dlg):
    dlg.chk_klarf.setChecked(False)
    assert dlg.run_export(sync=True) is None
    assert "Nothing is selected" in dlg.error_text()
