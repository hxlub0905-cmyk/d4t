# d4t report export — authored 2026-07-28 (M5-1).
"""批次結果報表：CSV（feature vector）＋ Excel（給人看的三頁報告）。

三個對外函式：

- :func:`summarize`   純資料統計 —— Excel 的兩頁與 CLI 都吃它，唯一的真相來源。
- :func:`write_csv`   utf-8-sig CSV（Excel 直接雙擊不亂碼），欄位與
  :meth:`d4t.core.store.RunStore.export_csv` 相同，方便兩邊互換。
- :func:`write_excel` openpyxl 活頁簿，三張工作表：

  ============  ==========================================================
  「摘要」       總數 / bin 分佈 / 分數統計 / recipe 資訊；有 ground truth
                時再加 **抓漏率、誤殺率、正確率** 與 2×2 混淆矩陣。
  「明細」       CSV 那張表，凍結標題列 + 自動篩選。
  「特徵統計」   每個特徵在整個 lot 的 最小 / 中位數 / 最大 / 標準差。
  ============  ==========================================================

名詞（半導體 ADC 的講法，別搞混）::

    抓漏 (miss)        真缺陷被判成 nuisance —— 漏掉了，最嚴重
    誤殺 (false alarm) nuisance 被判成真缺陷 —— 白花人力去 review

    抓漏率 = 抓漏數 / 真缺陷總數
    誤殺率 = 誤殺數 / nuisance 總數
    正確率 = 判對的顆數 / 有標註且有判定的顆數

``results`` 一律是 :func:`d4t.core.pipeline.result_to_json_dict` 形狀的
dict list（``defect_id`` / ``ok`` / ``error`` / ``score`` / ``bin`` /
``features``）。檔案寫入一律 atomic（``.tmp`` + :func:`os.replace`）。
"""
from __future__ import annotations

import csv
import math
import os
import statistics
from typing import (Any, Dict, Iterable, List, Optional, Sequence,
                    Set, Tuple)

from .klarf_out import ExportError

__all__ = ["summarize", "write_csv", "write_excel",
           "feature_keys", "BASE_COLUMNS"]

#: 明細表最前面的固定欄位（其後接排序過的特徵欄）。
BASE_COLUMNS = ("defect_id", "ok", "error", "score", "bin")

#: bin 為 None（沒判定）時在 bin 分佈裡使用的 key。
UNBINNED_KEY = "unbinned"


# ---------------------------------------------------------------------------
# 小工具
# ---------------------------------------------------------------------------
def _finite(v: Any) -> Optional[float]:
    """float 化；None / nan / inf / 不可轉 → None。"""
    if v is None or isinstance(v, bool):
        return None if v is None else float(v)
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or math.isinf(f):
        return None
    return f


def _atomic_replace(tmp: str, path: str) -> str:
    os.replace(tmp, path)
    return path


def _ensure_parent(path: str) -> None:
    parent = os.path.dirname(os.path.abspath(str(path)))
    if parent:
        os.makedirs(parent, exist_ok=True)


def feature_keys(results: Sequence[Dict[str, Any]]) -> List[str]:
    """所有結果的特徵 key 聯集（排序）—— 明細表與特徵統計的欄位順序。"""
    keys: Set[str] = set()
    for r in results or ():
        keys.update((r.get("features") or {}).keys())
    return sorted(keys)


def _gt_is_real(value: Any) -> Optional[bool]:
    """ground truth 的一筆 → 是不是真缺陷。

    接受三種寫法：``True`` / ``{"is_real": True, ...}``（``make_sample`` 的
    ``ground_truth.json``）/ 字串（``"1"`` ``"true"`` ``"real"`` …）。
    看不懂就回 None（該顆不列入統計，而不是猜）。
    """
    if isinstance(value, dict):
        for k in ("is_real", "real", "label", "truth"):
            if k in value:
                return _gt_is_real(value[k])
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        s = value.strip().lower()
        if s in ("1", "true", "yes", "y", "t", "real", "defect"):
            return True
        if s in ("0", "false", "no", "n", "f", "nuisance"):
            return False
    return None


def _lookup_gt(ground_truth: Dict[Any, Any], defect_id: Any) -> Optional[bool]:
    if defect_id in ground_truth:
        return _gt_is_real(ground_truth[defect_id])
    sid = str(defect_id)
    if sid in ground_truth:
        return _gt_is_real(ground_truth[sid])
    try:
        nid = str(int(sid))
    except (TypeError, ValueError):
        return None
    if nid in ground_truth:
        return _gt_is_real(ground_truth[nid])
    return None


def _stats(vals: Sequence[float]) -> Dict[str, Optional[float]]:
    if not vals:
        return {"n": 0, "min": None, "median": None, "max": None, "std": None}
    return {
        "n": len(vals),
        "min": float(min(vals)),
        "median": float(statistics.median(vals)),
        "max": float(max(vals)),
        "std": float(statistics.pstdev(vals)) if len(vals) > 1 else 0.0,
    }


# ---------------------------------------------------------------------------
# summarize —— 兩張工作表與 CLI 共用的純資料函式
# ---------------------------------------------------------------------------
def summarize(results: Sequence[Dict[str, Any]], *,
              ground_truth: Optional[Dict[Any, Any]] = None,
              positive_bins: Optional[Iterable[int]] = None) -> Dict[str, Any]:
    """把一批結果整理成報表要的所有數字（純資料，不碰檔案）。

    ``positive_bins``：哪些 bin 代表「判定為真缺陷」。預設 ``None`` =
    「bin != 0」（引擎的慣例是 below→0、above→1）。

    回傳::

        {
          "n_total", "n_ok", "n_fail", "n_scored",
          "bin_counts": {"0": 4, "1": 3, "未判定": 1},   # key 是字串
          "score_min", "score_median", "score_max",
          "features": {特徵名: {"n","min","median","max","std"}},
          "ground_truth": None 或 {
              "n_labelled", "n_evaluated", "tp", "fp", "tn", "fn",
              "miss_rate", "false_alarm_rate", "accuracy",
              "positive_bins"},
        }
    """
    results = list(results or [])
    n_total = len(results)
    n_ok = sum(1 for r in results if r.get("ok"))

    scores: List[float] = []
    bin_counts: Dict[str, int] = {}
    for r in results:
        s = _finite(r.get("score"))
        if s is not None:
            scores.append(s)
        b = r.get("bin")
        key = UNBINNED_KEY if b is None else str(int(b))
        bin_counts[key] = bin_counts.get(key, 0) + 1

    feats: Dict[str, List[float]] = {k: [] for k in feature_keys(results)}
    for r in results:
        for k, v in (r.get("features") or {}).items():
            f = _finite(v)
            if f is not None:
                feats[k].append(f)

    out: Dict[str, Any] = {
        "n_total": n_total,
        "n_ok": n_ok,
        "n_fail": n_total - n_ok,
        "n_scored": len(scores),
        "bin_counts": bin_counts,
        "score_min": min(scores) if scores else None,
        "score_median": statistics.median(scores) if scores else None,
        "score_max": max(scores) if scores else None,
        "features": {k: _stats(v) for k, v in sorted(feats.items())},
        "ground_truth": None,
    }
    if ground_truth:
        out["ground_truth"] = _confusion(results, ground_truth, positive_bins)
        out["bin_purity"] = _bin_purity(results, ground_truth)
    return out


def _bin_purity(results: Sequence[Dict[str, Any]],
                ground_truth: Dict[Any, Any]) -> List[Dict[str, Any]]:
    """**每一個 bin 裡有幾顆是真的**（F22-UI）—— 多類別唯一量得出來的東西。

    為什麼不是「多類別的正確率」
    ----------------------------
    正確率要先知道「這一顆應該落在哪一個 bin」，而那需要把 ground truth 標到
    **類別**。手上的資料（包含合成的那幾份）只標了 ``is_real`` ——
    一個二元的旗標。硬要算多類別正確率就得先假造一份對照，而那不是量出來的。

    純度不需要那個假設：它問的是「**我判進這一類的，有幾顆真的是缺陷**」，
    而那正是調規則的人一條一條在看的東西。一條規則太寬 → 那個 bin 的純度掉；
    太窄 → 那個 bin 的顆數掉。兩件事都看得見。

    ⚠ **`bin 0` 的純度要反過來讀**：它是「都沒對上」那一格，所以那裡的
    「真缺陷」是**漏抓**。所以每一列都同時給 ``n_real`` 與 ``n_nuisance``，
    不只給一個比例 —— 一個數字答不出「這一格好不好」，因為好壞取決於那一格
    本來要收什麼。

    回傳一串（照 bin 由大到小），每一項::

        {"bin": 3, "n": 11, "n_real": 11, "n_nuisance": 0,
         "n_unlabelled": 0, "purity": 1.0}

    ``purity`` 在整格都沒標註時是 ``None``（不是 0 —— 沒有分母不等於零純度，
    跟 CLI 的 `_pct` 是同一條規矩）。
    """
    buckets: Dict[Any, Dict[str, int]] = {}
    for r in results:
        b = r.get("bin")
        key = UNBINNED_KEY if b is None else int(b)
        row = buckets.setdefault(key, {"n": 0, "n_real": 0, "n_nuisance": 0,
                                       "n_unlabelled": 0})
        row["n"] += 1
        truth = _lookup_gt(ground_truth, r.get("defect_id"))
        if truth is None:
            row["n_unlabelled"] += 1
        elif truth:
            row["n_real"] += 1
        else:
            row["n_nuisance"] += 1

    def _order(k: Any) -> Tuple[int, int]:
        # 沒判定的那一格排最後（它不是一個 bin）
        return (1, 0) if k == UNBINNED_KEY else (0, -int(k))

    out: List[Dict[str, Any]] = []
    for key in sorted(buckets, key=_order):
        row = buckets[key]
        labelled = row["n_real"] + row["n_nuisance"]
        out.append({
            "bin": key,
            "n": row["n"],
            "n_real": row["n_real"],
            "n_nuisance": row["n_nuisance"],
            "n_unlabelled": row["n_unlabelled"],
            "purity": (row["n_real"] / labelled) if labelled else None,
        })
    return out


def _confusion(results: Sequence[Dict[str, Any]], ground_truth: Dict[Any, Any],
               positive_bins: Optional[Iterable[int]]) -> Dict[str, Any]:
    """抓漏 / 誤殺 / 正確率 + 2×2 混淆矩陣。

    判定為真缺陷 = ``bin in positive_bins``（預設 ``bin != 0``）。
    bin 是 None（沒判定）或 ground truth 沒標註的顆數不列入矩陣，
    另外記在 ``n_unbinned`` / ``n_unlabelled``。
    """
    pos: Optional[Set[int]] = (None if positive_bins is None
                               else {int(b) for b in positive_bins})
    tp = fp = tn = fn = 0
    n_labelled = 0
    n_unbinned = 0
    for r in results:
        truth = _lookup_gt(ground_truth, r.get("defect_id"))
        if truth is None:
            continue
        n_labelled += 1
        b = r.get("bin")
        if b is None:
            n_unbinned += 1
            continue
        b = int(b)
        pred = (b != 0) if pos is None else (b in pos)
        if truth and pred:
            tp += 1
        elif truth and not pred:
            fn += 1
        elif (not truth) and pred:
            fp += 1
        else:
            tn += 1

    n_eval = tp + fp + tn + fn
    n_real = tp + fn
    n_nuis = fp + tn
    return {
        "n_labelled": n_labelled,
        "n_unlabelled": len(results) - n_labelled,
        "n_unbinned": n_unbinned,
        "n_evaluated": n_eval,
        "n_real": n_real,
        "n_nuisance": n_nuis,
        "tp": tp, "fp": fp, "tn": tn, "fn": fn,
        # 抓漏率：真缺陷被判成 nuisance 的比例
        "miss_rate": (fn / n_real) if n_real else None,
        # 誤殺率：nuisance 被判成真缺陷的比例
        "false_alarm_rate": (fp / n_nuis) if n_nuis else None,
        "accuracy": ((tp + tn) / n_eval) if n_eval else None,
        "positive_bins": (None if pos is None else sorted(pos)),
    }


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def _detail_rows(results: Sequence[Dict[str, Any]], keys: Sequence[str]
                 ) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for r in results:
        feats = r.get("features") or {}
        row: List[Any] = [
            r.get("defect_id", ""),
            1 if r.get("ok") else 0,
            "" if r.get("error") is None else r.get("error"),
            "" if _finite(r.get("score")) is None else _finite(r.get("score")),
            "" if r.get("bin") is None else int(r.get("bin")),
        ]
        for k in keys:
            v = _finite(feats.get(k))
            row.append("" if v is None else v)
        rows.append(row)
    return rows


def write_csv(results: Sequence[Dict[str, Any]], path: str, *,
              include_features: bool = True) -> str:
    """匯出明細 CSV（**utf-8-sig**，Excel 雙擊直接開不亂碼），atomic 寫入。

    欄位：``defect_id, ok, error, score, bin`` 後面接**排序過的特徵 key
    聯集**（某顆沒有的特徵留空）—— 這張表就是 feature vector，可直接餵 ML。
    ``include_features=False`` 只輸出前五欄。
    """
    results = list(results or [])
    path = str(path)
    _ensure_parent(path)
    keys = feature_keys(results) if include_features else []
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(list(BASE_COLUMNS) + list(keys))
        for row in _detail_rows(results, keys):
            w.writerow(row)
    return _atomic_replace(tmp, path)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
_SHEET_SUMMARY = "Summary"
_SHEET_DETAIL = "Details"
_SHEET_FEATURES = "Feature stats"

_NUM_FMT = "0.0000"
_PCT_FMT = "0.00%"


def _recipe_info(recipe: Any) -> List[Sequence[Any]]:
    """Recipe 物件 / recipe JSON dict → 摘要頁要顯示的幾列。"""
    if recipe is None:
        return []
    if hasattr(recipe, "to_json_dict"):
        rd = recipe.to_json_dict()
    elif isinstance(recipe, dict):
        rd = recipe
    else:
        return [("recipe", str(recipe))]
    score = rd.get("score") or {}
    rows: List[Sequence[Any]] = [("Recipe name", rd.get("recipe_id", ""))]
    if rd.get("description"):
        rows.append(("Description", rd.get("description")))
    if score.get("expr") is not None:
        rows.append(("Score expression", str(score.get("expr"))))
    if score.get("threshold") is not None:
        rows.append(("Threshold", _finite(score.get("threshold"))))
    if score.get("bins"):
        rows.append(("Bin mapping", ", ".join(
            "{}={}".format(k, v) for k, v in sorted(score["bins"].items()))))
    if rd.get("nodes") is not None:
        try:
            rows.append(("Step count", len(rd["nodes"])))
        except TypeError:
            pass
    return rows


def _autosize(ws, widths: Sequence[int]) -> None:
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(8, min(60, w))


def _text_width(rows: Iterable[Sequence[Any]], n_cols: int) -> List[int]:
    widths = [10] * n_cols
    for row in rows:
        for i, v in enumerate(row[:n_cols]):
            widths[i] = max(widths[i], len(str("" if v is None else v)) + 2)
    return widths


def write_excel(results: Sequence[Dict[str, Any]], path: str, *,
                ground_truth: Optional[Dict[Any, Any]] = None,
                recipe: Any = None,
                positive_bins: Optional[Iterable[int]] = None) -> str:
    """匯出 Excel 報表（三張工作表），atomic 寫入。回傳寫入路徑。

    - 「摘要」：總數、bin 分佈、分數統計、recipe 資訊；給了 ``ground_truth``
      再加抓漏率 / 誤殺率 / 正確率與 2×2 混淆矩陣。
    - 「明細」：與 :func:`write_csv` 相同的表，凍結標題列 + 自動篩選。
    - 「特徵統計」：每個特徵的 筆數 / 最小 / 中位數 / 最大 / 標準差。

    ``ground_truth`` 可用 ``{defect_id: True}`` 或
    ``{defect_id: {"is_real": True, "type": "..."}}``（後者就是
    ``tools/make_sample.py`` 產出的 ``ground_truth.json``）。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:      # pragma: no cover — requirements 已含 openpyxl
        raise ExportError(
            "Exporting to Excel needs the openpyxl package, which is not "
            "installed on this machine ({}).\nRun: pip install openpyxl, or "
            "export to CSV instead.".format(e)) from None

    results = list(results or [])
    path = str(path)
    _ensure_parent(path)
    s = summarize(results, ground_truth=ground_truth, positive_bins=positive_bins)
    keys = feature_keys(results)

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="EDEFF2")

    wb = Workbook()

    # ---------------- 摘要 ----------------
    ws = wb.active
    ws.title = _SHEET_SUMMARY
    rows: List[Sequence[Any]] = [("Item", "Value"), ("__section__", "Overall")]
    rows += [
        ("Total defects", s["n_total"]),
        ("Succeeded", s["n_ok"]),
        ("Failed", s["n_fail"]),
        ("Scored", s["n_scored"]),
    ]
    rows.append(("__section__", "Bin distribution"))
    for k in sorted(s["bin_counts"], key=lambda x: (x == UNBINNED_KEY, x)):
        rows.append(("bin {}".format(k), s["bin_counts"][k]))
    rows.append(("__section__", "Score"))
    rows += [
        ("Minimum", s["score_min"]),
        ("Median", s["score_median"]),
        ("Maximum", s["score_max"]),
    ]
    rinfo = _recipe_info(recipe)
    if rinfo:
        rows.append(("__section__", "recipe"))
        rows += rinfo
    gt = s["ground_truth"]
    if gt:
        rows.append(("__section__", "Against ground truth"))
        rows += [
            ("Labelled defects", gt["n_labelled"]),
            ("Actually evaluated", gt["n_evaluated"]),
            ("Miss rate (real defect called nuisance)", gt["miss_rate"]),
            ("False alarm rate (nuisance called real defect)", gt["false_alarm_rate"]),
            ("Accuracy", gt["accuracy"]),
            ("__section__", "Confusion matrix"),
            ("", "Called: real defect", "Called: nuisance"),
            ("Actual: real defect", gt["tp"], gt["fn"]),
            ("Actual: nuisance", gt["fp"], gt["tn"]),
        ]

    r_i = 0
    pct_rows = set()
    for row in rows:
        r_i += 1
        if row[0] == "__section__":
            ws.cell(row=r_i, column=1, value=str(row[1])).font = bold
            ws.cell(row=r_i, column=1).fill = head_fill
            continue
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r_i, column=c_i, value=v)
            if r_i == 1 or (c_i == 1 and isinstance(v, str)
                            and v.startswith("Actual:")):
                cell.font = bold
        label = str(row[0])
        if label.startswith(("Miss rate", "False alarm rate", "Accuracy")):
            pct_rows.add(r_i)
    for r in pct_rows:
        ws.cell(row=r, column=2).number_format = _PCT_FMT
    for label in ("Minimum", "Median", "Maximum", "Threshold"):
        for r in range(1, r_i + 1):
            if ws.cell(row=r, column=1).value == label:
                ws.cell(row=r, column=2).number_format = _NUM_FMT
    ws.cell(row=1, column=1).fill = head_fill
    ws.cell(row=1, column=2).fill = head_fill
    _autosize(ws, _text_width(rows, 3))
    ws.freeze_panes = "A2"

    # ---------------- 明細 ----------------
    ws = wb.create_sheet(_SHEET_DETAIL)
    header = list(BASE_COLUMNS) + list(keys)
    ws.append(header)
    for c_i in range(1, len(header) + 1):
        c = ws.cell(row=1, column=c_i)
        c.font = bold
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")
    detail = _detail_rows(results, keys)
    for row in detail:
        ws.append(["" if v is None else v for v in row])
    n_rows = len(detail) + 1
    for c_i in range(4, len(header) + 1):        # score 之後全是數值欄
        if c_i == 5:                             # bin 是整數
            continue
        for r_i in range(2, n_rows + 1):
            ws.cell(row=r_i, column=c_i).number_format = _NUM_FMT
    ws.freeze_panes = "A2"
    if n_rows >= 1:
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = "A1:{}{}".format(
            get_column_letter(len(header)), max(n_rows, 1))
    _autosize(ws, _text_width([header] + detail, len(header)))

    # ---------------- 特徵統計 ----------------
    ws = wb.create_sheet(_SHEET_FEATURES)
    fheader = ["Feature", "Count", "Minimum", "Median", "Maximum", "Std dev"]
    ws.append(fheader)
    for c_i in range(1, len(fheader) + 1):
        c = ws.cell(row=1, column=c_i)
        c.font = bold
        c.fill = head_fill
    frows: List[Sequence[Any]] = []
    for name, st in s["features"].items():
        frows.append((name, st["n"], st["min"], st["median"], st["max"], st["std"]))
    for row in frows:
        ws.append(list(row))
    for r_i in range(2, len(frows) + 2):
        for c_i in range(3, 7):
            ws.cell(row=r_i, column=c_i).number_format = _NUM_FMT
    ws.freeze_panes = "A2"
    _autosize(ws, _text_width([fheader] + frows, len(fheader)))

    tmp = path + ".tmp"
    wb.save(tmp)
    return _atomic_replace(tmp, path)
